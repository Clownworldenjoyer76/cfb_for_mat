#!/usr/bin/env python3
from __future__ import annotations

import argparse
import importlib.util
import math
import os
import re
import sys
import tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from types import ModuleType
from typing import Any

import numpy as np
import pandas as pd

SCRIPT_DIR = Path(__file__).resolve().parent
CFB_ROOT = SCRIPT_DIR.parents[1]

PICKS_PATH = CFB_ROOT / "scripts" / "03_picks" / "picks.py"
GRADE_PATH = CFB_ROOT / "scripts" / "04_final_results" / "grade_picks.py"
SCORES_PATH = CFB_ROOT / "scripts" / "04_final_results" / "pull_final_scores.py"
MARKETS_PATH = CFB_ROOT / "config" / "markets.yaml"

FILE_RE = re.compile(r"^week_(\d+)_CFB_selected\.csv$")


def load_module(name: str, path: Path) -> ModuleType:
    spec = importlib.util.spec_from_file_location(name, path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Unable to import {path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


def clean(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = str(value).strip()
    if text.casefold() in {"", "nan", "none", "null", "<na>", "nat"}:
        return ""
    return text


def number(value: Any) -> float | None:
    text = clean(value)
    if not text:
        return None
    try:
        value = float(text)
    except (TypeError, ValueError):
        return None
    return value if math.isfinite(value) else None


def flag(value: Any) -> bool:
    return clean(value).casefold() in {"1", "1.0", "true", "yes", "y"}


def normalize_game_id(value: Any) -> str:
    text = clean(value)
    return text[:-2] if re.fullmatch(r"\d+\.0", text) else text


def simulate_filters(
    files: list[Path],
    picks: ModuleType,
    config: dict[str, Any],
) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []

    with tempfile.TemporaryDirectory(prefix="cfb_bt_") as tmp:
        tmp_path = Path(tmp)

        for path in files:
            out = tmp_path / picks.output_name(path)
            frame = picks.process_file(path, out, config)
            frame["source_file"] = path.name
            frames.append(frame)

    combined = pd.concat(frames, ignore_index=True)
    combined["game_id"] = combined["game_id"].map(normalize_game_id)
    return combined


def fetch_scores(
    game_ids: list[str],
    scores: ModuleType,
    workers: int,
) -> dict[str, dict[str, Any]]:
    results: dict[str, dict[str, Any]] = {}

    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {
            executor.submit(scores.get_score_and_status, game_id): game_id
            for game_id in game_ids
        }

        done = 0
        total = len(futures)

        for future in as_completed(futures):
            game_id = futures[future]
            try:
                results[game_id] = future.result()
            except Exception as exc:
                results[game_id] = {
                    "away_score": "",
                    "home_score": "",
                    "completed": 0,
                    "status": "",
                    "fetch_error": str(exc),
                }

            done += 1
            if done % 50 == 0 or done == total:
                print(f"final_scores={done}/{total}")

    return results


def grade_one(
    market: str,
    selection: str,
    line: float | None,
    home_score: float | None,
    away_score: float | None,
    completed: bool,
    voided: bool,
    grade: ModuleType,
) -> str:
    if market == "moneyline":
        return grade.grade_moneyline(
            True,
            selection,
            home_score,
            away_score,
            completed,
            voided,
        )

    if market == "spread":
        return grade.grade_spread(
            True,
            selection,
            line,
            home_score,
            away_score,
            completed,
            voided,
        )

    return grade.grade_total(
        True,
        selection,
        line,
        home_score,
        away_score,
        completed,
        voided,
    )


def bucket_ev(x: Any) -> str:
    x = number(x)
    if x is None:
        return "NA"
    if x < 0:
        return "<0"
    if x < .02:
        return "0-.0199"
    if x < .04:
        return ".02-.0399"
    if x < .06:
        return ".04-.0599"
    if x < .10:
        return ".06-.0999"
    if x < .15:
        return ".10-.1499"
    if x < .25:
        return ".15-.2499"
    return ".25+"


def bucket_kelly(x: Any) -> str:
    x = number(x)
    if x is None:
        return "NA"
    if x <= 0:
        return "0"
    if x < .005:
        return "0-.49%"
    if x < .01:
        return ".50-.99%"
    if x < .02:
        return "1-1.99%"
    if x < .03:
        return "2-2.99%"
    if x < .05:
        return "3-4.99%"
    if x < .10:
        return "5-9.99%"
    return "10%+"


def bucket_prob(x: Any) -> str:
    x = number(x)
    if x is None:
        return "NA"
    if x < .50:
        return "<50%"
    if x < .55:
        return "50-54.99%"
    if x < .60:
        return "55-59.99%"
    if x < .65:
        return "60-64.99%"
    if x < .70:
        return "65-69.99%"
    if x < .80:
        return "70-79.99%"
    if x < .90:
        return "80-89.99%"
    return "90%+"


def bucket_edge(x: Any) -> str:
    x = number(x)
    if x is None:
        return "NA"
    if x < 0:
        return "<0"
    if x < .02:
        return "0-.0199"
    if x < .04:
        return ".02-.0399"
    if x < .06:
        return ".04-.0599"
    if x < .10:
        return ".06-.0999"
    if x < .15:
        return ".10-.1499"
    return ".15+"


def bucket_odds(x: Any) -> str:
    x = number(x)
    if x is None:
        return "NA"
    if x <= -500:
        return "<=-500"
    if x <= -300:
        return "-499:-300"
    if x <= -200:
        return "-299:-200"
    if x <= -150:
        return "-199:-150"
    if x <= -110:
        return "-149:-110"
    if x <= -101:
        return "-109:-101"
    if 100 <= x <= 149:
        return "+100:+149"
    if x <= 199:
        return "+150:+199"
    if x <= 299:
        return "+200:+299"
    if x <= 499:
        return "+300:+499"
    if x >= 500:
        return "+500+"
    return "INVALID"


def make_ledger(
    games: pd.DataFrame,
    score_map: dict[str, dict[str, Any]],
    picks: ModuleType,
    grade: ModuleType,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []

    for _, game in games.iterrows():
        game_id = normalize_game_id(game.get("game_id"))
        final = score_map.get(game_id, {})

        home_score = number(final.get("home_score"))
        away_score = number(final.get("away_score"))
        completed = flag(final.get("completed"))
        status = clean(final.get("status"))
        voided = grade.is_void_status(status)

        for market_name, spec in picks.MARKETS.items():
            out_prefix = spec["output_prefix"]
            chosen = clean(game.get(f"{out_prefix}_selection")).upper()
            market_selected = flag(game.get(f"{out_prefix}_selected"))

            for side_name, (candidate_prefix, selection) in spec["sides"].items():
                available = flag(game.get(f"{candidate_prefix}_available"))
                odds = number(game.get(f"{candidate_prefix}_odds_american"))
                model_prob = number(game.get(f"{candidate_prefix}_model_probability"))
                implied_prob = number(game.get(f"{candidate_prefix}_implied_probability"))
                edge = number(game.get(f"{candidate_prefix}_edge"))
                ev = number(game.get(f"{candidate_prefix}_ev"))
                full_kelly = number(game.get(f"{candidate_prefix}_full_kelly"))
                candidate_kelly = number(game.get(f"{candidate_prefix}_kelly"))
                line = (
                    number(game.get(f"{candidate_prefix}_line"))
                    if market_name in {"spread", "total"}
                    else None
                )

                selected = int(
                    available
                    and market_selected
                    and chosen == selection
                )

                if available:
                    result = grade_one(
                        market_name,
                        selection,
                        line,
                        home_score,
                        away_score,
                        completed,
                        voided,
                        grade,
                    )
                    profit = grade.profit_for_grade(result, odds)
                else:
                    result = "NO_MARKET"
                    profit = None

                rows.append(
                    {
                        "season": game.get("season"),
                        "season_type": game.get("season_type"),
                        "week": game.get("week"),
                        "game_id": game_id,
                        "game_date": game.get("game_date"),
                        "away_team": game.get("away_team"),
                        "home_team": game.get("home_team"),
                        "market": market_name,
                        "side": side_name.upper(),
                        "selection": selection,
                        "candidate_available": int(available),
                        "selected": selected,
                        "odds_american": odds,
                        "line": line,
                        "model_probability": model_prob,
                        "implied_probability": implied_prob,
                        "edge": edge,
                        "ev": ev,
                        "full_kelly": full_kelly,
                        "kelly": candidate_kelly,
                        "historical_source_bookmaker": game.get(
                            "historical_source_bookmaker", ""
                        ),
                        "historical_provider_proxy": game.get(
                            "historical_provider_proxy", ""
                        ),
                        "final_status": status,
                        "final_completed": int(completed),
                        "away_score": away_score,
                        "home_score": home_score,
                        "grade": result,
                        "profit_units": profit,
                        "selected_profit_units": (
                            profit if selected and profit is not None else 0.0
                        ),
                        "ev_range": bucket_ev(ev),
                        "kelly_range": bucket_kelly(candidate_kelly),
                        "prob_range": bucket_prob(model_prob),
                        "odds_range": bucket_odds(odds),
                        "edge_range": bucket_edge(edge),
                        "source_file": game.get("source_file", ""),
                    }
                )

    return pd.DataFrame(rows)


def summarize(frame: pd.DataFrame, groups: list[str]) -> pd.DataFrame:
    def one(group: pd.DataFrame) -> dict[str, Any]:
        grades = group["grade"].astype(str)
        wins = int(grades.eq("WIN").sum())
        losses = int(grades.eq("LOSS").sum())
        pushes = int(grades.eq("PUSH").sum())
        graded = wins + losses + pushes
        decisions = wins + losses

        profit = pd.to_numeric(
            group.loc[grades.isin(["WIN", "LOSS", "PUSH"]), "profit_units"],
            errors="coerce",
        ).sum()

        avg_model = pd.to_numeric(
            group.loc[grades.isin(["WIN", "LOSS"]), "model_probability"],
            errors="coerce",
        ).mean()

        actual = wins / decisions if decisions else np.nan

        return {
            "bets": graded,
            "wins": wins,
            "losses": losses,
            "pushes": pushes,
            "win_rate": actual,
            "avg_odds": pd.to_numeric(group["odds_american"], errors="coerce").mean(),
            "avg_model_probability": avg_model,
            "calibration_gap": (
                actual - avg_model
                if decisions and not pd.isna(avg_model)
                else np.nan
            ),
            "avg_edge": pd.to_numeric(group["edge"], errors="coerce").mean(),
            "avg_ev": pd.to_numeric(group["ev"], errors="coerce").mean(),
            "avg_kelly": pd.to_numeric(group["kelly"], errors="coerce").mean(),
            "net_units": float(profit),
            "roi": float(profit) / graded if graded else np.nan,
        }

    if not groups:
        return pd.DataFrame([one(frame)])

    rows: list[dict[str, Any]] = []

    for keys, group in frame.groupby(groups, dropna=False, sort=False):
        if not isinstance(keys, tuple):
            keys = (keys,)
        row = {name: value for name, value in zip(groups, keys)}
        row.update(one(group))
        rows.append(row)

    return pd.DataFrame(rows)


def write_workbook(
    path: Path,
    ledger: pd.DataFrame,
) -> None:
    selected = ledger[
        ledger["candidate_available"].eq(1)
        & ledger["selected"].eq(1)
    ].copy()

    candidates = ledger[
        ledger["candidate_available"].eq(1)
    ].copy()

    tables = {
        "selected_overall": summarize(selected, []),
        "selected_market": summarize(selected, ["market"]),
        "selected_week": summarize(selected, ["week"]),
        "selected_ev": summarize(selected, ["ev_range"]),
        "selected_kelly": summarize(selected, ["kelly_range"]),
        "selected_prob": summarize(selected, ["prob_range"]),
        "selected_odds": summarize(selected, ["odds_range"]),
        "selected_edge": summarize(selected, ["edge_range"]),
        "market_ev": summarize(selected, ["market", "ev_range"]),
        "market_kelly": summarize(selected, ["market", "kelly_range"]),
        "market_prob": summarize(selected, ["market", "prob_range"]),
        "market_odds": summarize(selected, ["market", "odds_range"]),
        "market_edge": summarize(selected, ["market", "edge_range"]),
        "ev_kelly": summarize(selected, ["ev_range", "kelly_range"]),
        "ev_prob": summarize(selected, ["ev_range", "prob_range"]),
        "ev_odds": summarize(selected, ["ev_range", "odds_range"]),
        "ev_edge": summarize(selected, ["ev_range", "edge_range"]),
        "kelly_prob": summarize(selected, ["kelly_range", "prob_range"]),
        "kelly_odds": summarize(selected, ["kelly_range", "odds_range"]),
        "kelly_edge": summarize(selected, ["kelly_range", "edge_range"]),
        "prob_odds": summarize(selected, ["prob_range", "odds_range"]),
        "prob_edge": summarize(selected, ["prob_range", "edge_range"]),
        "odds_edge": summarize(selected, ["odds_range", "edge_range"]),
        "all_candidates": summarize(candidates, ["market"]),
        "candidate_ev": summarize(candidates, ["ev_range"]),
        "candidate_kelly": summarize(candidates, ["kelly_range"]),
        "candidate_prob": summarize(candidates, ["prob_range"]),
        "candidate_odds": summarize(candidates, ["odds_range"]),
        "candidate_edge": summarize(candidates, ["edge_range"]),
    }

    tmp = path.with_suffix(".tmp.xlsx")

    with pd.ExcelWriter(tmp, engine="openpyxl") as writer:
        for sheet, table in tables.items():
            table.to_excel(writer, sheet_name=sheet[:31], index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = load_workbook(tmp)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for col in range(1, ws.max_column + 1):
            max_len = 10
            for row in range(1, min(ws.max_row, 3000) + 1):
                value = ws.cell(row=row, column=col).value
                if value is not None:
                    max_len = max(max_len, len(str(value)) + 2)
            ws.column_dimensions[get_column_letter(col)].width = min(max_len, 34)

    wb.save(tmp)
    os.replace(tmp, path)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--season", type=int, default=2025)
    parser.add_argument("--workers", type=int, default=8)
    args = parser.parse_args()

    candidate_dir = (
        CFB_ROOT
        / "05_backtest"
        / "input"
        / str(args.season)
        / "candidates"
    )

    output_dir = (
        CFB_ROOT
        / "05_backtest"
        / "output"
        / str(args.season)
    )

    files = sorted(
        path
        for path in candidate_dir.glob("week_*_CFB_selected.csv")
        if FILE_RE.fullmatch(path.name)
    )

    if not files:
        raise RuntimeError(f"No candidate files found in {candidate_dir}")

    picks = load_module("cfb_bt_picks", PICKS_PATH)
    grade = load_module("cfb_bt_grade", GRADE_PATH)
    scores = load_module("cfb_bt_scores", SCORES_PATH)

    config = picks.normalize_config(
        picks.load_yaml(MARKETS_PATH)
    )

    games = simulate_filters(files, picks, config)

    score_map = fetch_scores(
        games["game_id"].map(normalize_game_id).drop_duplicates().tolist(),
        scores,
        args.workers,
    )

    ledger = make_ledger(
        games,
        score_map,
        picks,
        grade,
    )

    output_dir.mkdir(parents=True, exist_ok=True)

    game_level = (
        output_dir
        / f"{args.season}_CFB_backtest_game_level.csv"
    )

    summary = (
        output_dir
        / f"{args.season}_CFB_backtest_summary.xlsx"
    )

    ledger.to_csv(game_level, index=False)
    write_workbook(summary, ledger)

    selected = ledger[
        ledger["candidate_available"].eq(1)
        & ledger["selected"].eq(1)
    ]

    overall = summarize(selected, [])

    print("backtest completed")
    print(f"games={games['game_id'].nunique()}")
    print(f"selected_bets={len(selected)}")

    if not overall.empty:
        print(f"wins={overall.iloc[0]['wins']}")
        print(f"losses={overall.iloc[0]['losses']}")
        print(f"net_units={overall.iloc[0]['net_units']}")
        print(f"roi={overall.iloc[0]['roi']}")

    print(f"game_level={game_level}")
    print(f"summary={summary}")
    print("status=success")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
